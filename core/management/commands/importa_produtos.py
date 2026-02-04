from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
import pandas as pd
from PIL import Image, ImageFilter
import io

class Command(BaseCommand):
    help = 'Importa produtos do Excel e envia imagens para o Oracle Object Storage'

    def add_arguments(self, parser):
        parser.add_argument(
            '--preview',
            action='store_true',
            help='Mostra o que será feito sem enviar nada'
        )

    def process_image(self, img_data):
        """
        
        """
        try:
            image = Image.open(io.BytesIO(img_data))
            
            # Converter para RGBA para garantir canal alfa
            if image.mode != 'RGBA':
                image = image.convert('RGBA')
            
            # LOG DEBUG: Tamanho original
            self.stdout.write(f"DEBUG: Imagem original size: {image.size}")

            new_width = int(image.width * 1)
            new_height = int(image.height * 1)
            
            # Redimensionar com alta qualidade (LANCZOS)
            image_resized = image.resize((new_width, new_height), Image.Resampling.LANCZOS)
            
            # Aplicar filtro de nitidez para tentar melhorar o upscale
            image_resized = image_resized.filter(ImageFilter.UnsharpMask(radius=2, percent=150, threshold=3))

            # Salvar direto (sem canvas)
            output = io.BytesIO()
            image_resized.save(output, format='PNG', optimize=True)
            return output.getvalue()
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Erro ao processar imagem: {e}"))
            return None

    def handle(self, *args, **options):
        PREVIEW = options['preview']

        # ========= CONFIGURAÇÕES =========
        EXCEL_PATH = "KOHLER.xlsx"
        COL_CODIGO = 4      # Coluna A
        COL_DESC = 7        # Coluna B
        BUCKET_NAME = "produtos"
        NAMESPACE = "grsxg5eatn7l"
        REGION = "sa-saopaulo-1"
        OUTPUT_CSV = "produtos_com_url.csv"
        # =================================

        self.stdout.write(self.style.WARNING(
            "MODO PREVIEW ATIVO" if PREVIEW else "MODO EXECUÇÃO REAL"
        ))

        # Inicializa cliente OCI apenas se for executar
        if not PREVIEW:
            import oci
            config = oci.config.from_file()
            object_storage = oci.object_storage.ObjectStorageClient(config)

        try:
            # Carregar o arquivo Excel
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"Erro: Arquivo {EXCEL_PATH} não encontrado."))
            return
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Erro inesperado ao carregar o arquivo Excel: {e}"))
            return

        produtos_processados = []

        # ===============================
        # LEITURA DAS IMAGENS DO EXCEL
        # ===============================
        self.stdout.write(
            f"Total de imagens encontradas no Excel: {len(ws._images)}"
        )

        for idx, img in enumerate(ws._images, start=1):
            anchor = img.anchor
            row_start = -1
            row_end = -1

            try:
                if isinstance(anchor, TwoCellAnchor):
                    row_start = anchor._from.row + 1
                    row_end = anchor.to.row + 1
                elif isinstance(anchor, OneCellAnchor):
                    row_start = anchor._from.row + 1
                    row_end = row_start
                else:
                    # Tenta ler _from se disponível
                    if hasattr(anchor, '_from'):
                        row_start = anchor._from.row + 1
                        row_end = row_start
                    else:
                        self.stdout.write(self.style.WARNING(f"Imagem {idx}: Tipo de âncora desconhecido ({type(anchor)})."))
                        continue
            except AttributeError:
                self.stdout.write(self.style.WARNING(f"Imagem {idx}: Erro ao ler propriedades da âncora."))
                continue

            if row_end < row_start:
                row_end = row_start

            # Coletar todos os códigos no intervalo vertical da imagem
            found_items = []
            for r in range(row_start, row_end + 1):
                c = ws.cell(row=r, column=COL_CODIGO).value
                d = ws.cell(row=r, column=COL_DESC).value
                if c:
                    found_items.append({'row': r, 'code': c, 'desc': d})


            if not found_items:
                self.stdout.write(
                    self.style.WARNING(f"Imagem {idx} (Linhas {row_start}-{row_end}): ignorada (sem código no range)")
                )
                continue

            # Processar cada código encontrado para esta imagem
            for item in found_items:
                codigo = item['code']
                descricao = item['desc']
                row_found = item['row']
                
                nome_arquivo = f"{codigo}.png"
                url = (
                    f"https://objectstorage.{REGION}.oraclecloud.com"
                    f"/n/{NAMESPACE}/b/{BUCKET_NAME}/o/produtos/{nome_arquivo}"
                )

                # ===============================
                # PREVIEW — O QUE SERÁ ENVIADO
                # ===============================
                self.stdout.write(
                    f"[{idx}] Produto: {codigo} | Linha: {row_found} (Range Img: {row_start}-{row_end})"
                )
                self.stdout.write(f"     Arquivo: {nome_arquivo}")
                self.stdout.write(f"     URL: {url}")

                # ===============================
                # ENVIO REAL (se não for preview)
                # ===============================
                if not PREVIEW:
                    try:
                        # Processar imagem antes do upload
                        processed_img_data = self.process_image(img._data())
                        
                        if processed_img_data:
                            object_storage.put_object(
                                namespace_name=NAMESPACE,
                                bucket_name=BUCKET_NAME,
                                object_name=f"produtos/{nome_arquivo}",
                                put_object_body=processed_img_data,
                                content_type="image/png"
                            )
                            self.stdout.write(self.style.SUCCESS(f"     Upload concluído"))
                        else:
                            self.stdout.write(self.style.ERROR(f"     Falha ao processar imagem para {codigo}"))

                    except Exception as e:
                        self.stdout.write(self.style.ERROR(f"     Erro no upload: {e}"))

                produtos_processados.append({
                    "codigo": codigo,
                    "descricao": descricao,
                    "prod_url": url
                })

        # ===============================
        # PRODUTOS PROCESSADOS (RESUMO)
        # ===============================
        self.stdout.write(
            self.style.WARNING(
                f"Total de produtos processados: {len(produtos_processados)}"
            )
        )

        # ===============================
        # GERA CSV FINAL
        # ===============================
        df = pd.DataFrame(produtos_processados)
        df.to_csv(OUTPUT_CSV, index=False)

        self.stdout.write(
            self.style.SUCCESS(f"CSV gerado: {OUTPUT_CSV}")
        )

        if PREVIEW:
            self.stdout.write(
                self.style.WARNING("Nenhum arquivo foi enviado (modo preview).")
            )
